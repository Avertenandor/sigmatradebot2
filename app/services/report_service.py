"""
Report service.

Generates Excel reports for users.
"""

import io
from datetime import datetime
from decimal import Decimal

import openpyxl
from loguru import logger
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.deposit import Deposit
from app.models.enums import TransactionStatus
from app.models.referral import Referral
from app.models.transaction import Transaction
from app.models.user import User
from app.repositories.referral_earning_repository import ReferralEarningRepository


class ReportService:
    """Service for generating user reports."""

    def __init__(self, session: AsyncSession) -> None:
        """Initialize report service."""
        self.session = session
        self.earning_repo = ReferralEarningRepository(session)

    async def generate_user_report(self, user_id: int) -> bytes:
        """
        Generate comprehensive Excel report for user.

        Args:
            user_id: User ID

        Returns:
            Bytes of the Excel file
        """
        # Fetch data
        user = await self._get_user(user_id)
        transactions = await self._get_transactions(user_id)
        deposits = await self._get_deposits(user_id)
        referrals = await self._get_referrals(user_id)
        earnings = await self.earning_repo.get_all_for_referrer(user_id)
        wallet_history = await self._get_wallet_history(user_id)

        # Create workbook
        wb = openpyxl.Workbook()
        
        # Sheet 1: General Info
        self._create_general_sheet(wb, user, deposits, earnings)
        
        # Sheet 2: Transactions
        self._create_transactions_sheet(wb, transactions)
        
        # Sheet 3: Deposits
        self._create_deposits_sheet(wb, deposits)
        
        # Sheet 4: Referrals
        self._create_referrals_sheet(wb, referrals)

        # Sheet 5: Wallet History
        self._create_wallet_history_sheet(wb, wallet_history)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    async def _get_user(self, user_id: int) -> User:
        stmt = select(User).where(User.id == user_id)
        result = await self.session.execute(stmt)
        return result.scalar_one()

    async def _get_transactions(self, user_id: int) -> list[Transaction]:
        stmt = select(Transaction).where(Transaction.user_id == user_id).order_by(Transaction.created_at.desc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_deposits(self, user_id: int) -> list[Deposit]:
        stmt = (
            select(Deposit)
            .where(Deposit.user_id == user_id)
            .options(selectinload(Deposit.deposit_version))
            .order_by(Deposit.created_at.desc())
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_referrals(self, user_id: int) -> list[Referral]:
        stmt = (
            select(Referral)
            .where(Referral.referrer_id == user_id)
            .options(selectinload(Referral.referral))
            .order_by(Referral.created_at.desc())
        )
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    async def _get_wallet_history(self, user_id: int) -> list:
        from app.models.user_wallet_history import UserWalletHistory
        stmt = select(UserWalletHistory).where(UserWalletHistory.user_id == user_id).order_by(UserWalletHistory.changed_at.desc())
        result = await self.session.execute(stmt)
        return list(result.scalars().all())

    def _apply_header_style(self, ws):
        """Apply professional style to the header row."""
        header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            # Cap width at 50 chars to avoid super wide columns
            ws.column_dimensions[column].width = min(adjusted_width, 50)

    def _apply_zebra_striping(self, ws):
        """Apply alternating row colors for readability."""
        fill_even = PatternFill(start_color="E9EFF7", end_color="E9EFF7", fill_type="solid")
        thin_border = Side(border_style="thin", color="D4D4D4")
        border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if row[0].row % 2 == 0:
                    cell.fill = fill_even

    def _create_general_sheet(self, wb, user: User, deposits: list[Deposit], earnings: list):
        ws = wb.active
        ws.title = "Общая информация"
        
        # Styles
        title_font = Font(bold=True, size=14, color="366092")
        label_font = Font(bold=True)
        
        ws.append(["ОТЧЕТ ПОЛЬЗОВАТЕЛЯ SIGMATRADE"])
        ws["A1"].font = title_font
        ws.merge_cells("A1:B1")
        
        data = [
            ("ID пользователя", user.id),
            ("Telegram ID", user.telegram_id),
            ("Username", f"@{user.username}" if user.username else "Не указан"),
            ("Кошелек", user.wallet_address),
            ("Дата регистрации", user.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("", ""),
            ("💰 Баланс", float(user.balance)),
            ("💵 Всего заработано", float(user.total_earned)),
            ("⏳ Ожидает выплаты", float(user.pending_earnings)),
            ("", ""),
            ("📊 Статистика депозитов", ""),
            ("Всего депозитов", len(deposits)),
            ("Активных депозитов", len([d for d in deposits if d.status == TransactionStatus.CONFIRMED.value and not d.is_roi_completed])),
            ("Общая сумма депозитов", float(sum(d.amount for d in deposits))),
            ("", ""),
            ("🎁 Реферальная статистика", ""),
            ("Всего начислений", len(earnings)),
            ("Общая сумма начислений", float(sum(e.amount for e in earnings))),
        ]

        for row in data:
            ws.append(row)

        # Style labels (first column)
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                cell.font = label_font

        # Formatting
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

    def _create_transactions_sheet(self, wb, transactions: list[Transaction]):
        ws = wb.create_sheet("История транзакций")
        
        headers = ["ID", "Дата", "Тип", "Сумма (USDT)", "Статус", "Описание", "TX Hash", "Баланс до", "Баланс после"]
        ws.append(headers)
        
        for tx in transactions:
            ws.append([
                tx.id,
                tx.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                tx.type,
                float(tx.amount),
                tx.status,
                tx.description,
                tx.tx_hash or "-",
                float(tx.balance_before),
                float(tx.balance_after)
            ])
            
        self._apply_header_style(ws)
        self._apply_zebra_striping(ws)
        self._adjust_column_widths(ws)

    def _create_deposits_sheet(self, wb, deposits: list[Deposit]):
        ws = wb.create_sheet("Депозиты")
        
        headers = ["ID", "Дата", "Уровень", "Сумма (USDT)", "Статус", "ROI Cap", "Выплачено", "Завершено", "Процент дохода", "TX Hash"]
        ws.append(headers)
        
        for dep in deposits:
            roi_percent = "N/A"
            if dep.deposit_version and dep.deposit_version.roi_percent:
                roi_percent = f"{float(dep.deposit_version.roi_percent)}%"
            
            ws.append([
                dep.id,
                dep.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                dep.level,
                float(dep.amount),
                dep.status,
                float(dep.roi_cap_amount),
                float(dep.roi_paid_amount),
                "Да" if dep.is_roi_completed else "Нет",
                roi_percent,
                dep.tx_hash or "-"
            ])
            
        self._apply_header_style(ws)
        self._apply_zebra_striping(ws)
        self._adjust_column_widths(ws)

    def _create_referrals_sheet(self, wb, referrals: list[Referral]):
        ws = wb.create_sheet("Рефералы")
        
        headers = ["ID", "Дата регистрации", "Уровень", "Пользователь (Username)", "Пользователь (ID)", "Заработано с него (USDT)"]
        ws.append(headers)
        
        for ref in referrals:
            username = "Не указан"
            telegram_id = "Неизвестно"
            
            if ref.referral:
                username = f"@{ref.referral.username}" if ref.referral.username else "Не указан"
                telegram_id = str(ref.referral.telegram_id)
            
            ws.append([
                ref.id,
                ref.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                ref.level,
                username,
                telegram_id,
                float(ref.total_earned)
            ])
            
        self._apply_header_style(ws)
        self._apply_zebra_striping(ws)
        self._adjust_column_widths(ws)

    def _create_wallet_history_sheet(self, wb, history: list):
        ws = wb.create_sheet("История смены кошелька")
        
        headers = ["Дата изменения", "Старый кошелек", "Новый кошелек"]
        ws.append(headers)
        
        for h in history:
            ws.append([
                h.changed_at.strftime("%Y-%m-%d %H:%M:%S"),
                h.old_wallet_address,
                h.new_wallet_address
            ])
            
        self._apply_header_style(ws)
        self._apply_zebra_striping(ws)
        self._adjust_column_widths(ws)
